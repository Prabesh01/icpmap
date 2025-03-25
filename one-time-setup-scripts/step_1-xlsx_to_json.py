import json

import openpyxl

from config import get_basedir


BASE_DIR=get_basedir()


def read_xlsx(source):
    dataframe = openpyxl.load_workbook(source)

    classes=[]
    rooms=[]
    teachers=[]

    sections={}

    for sheet in dataframe.sheetnames:
        sections[sheet]=[]
        dataframe1=dataframe[sheet]
        cont=False
        for row in range(0, dataframe1.max_row):
            i=0
            individual_class={"class":sheet}
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                val=col[row].value
                if val: val=str(val).strip()
                if not val:
                    i+=1
                    continue
                if val=="SUN":cont=True
                if not cont: continue
                if i==0: individual_class["day"]=val
                elif i==1:
                    individual_class["stime"],individual_class["etime"]=val.split(' - ')
                elif i==2:
                    individual_class["type"]=val
                elif i==6: individual_class["code"]=val
                elif i==7: individual_class["module"]=val
                elif i==8:
                    if not val in teachers: teachers.append(val)
                    individual_class["teacher"]=val
                elif i==9:
                    individual_class["sections"]=[x.strip() for x in val.split('+')]
                    for x in val.split('+'):
                        if not x.strip() in sections[sheet]: sections[sheet].append(x.strip())
                elif i==11:
                    room=val.replace(" ", "").split("-")[-1].strip()
                    if not room in rooms: rooms.append(room)
                    individual_class["room"]=room
                i+=1
            if individual_class!={"class":sheet}: classes.append(individual_class)
        sections[sheet].sort()

    data={"classes":classes, "sections": sections, "rooms":rooms, "teachers":teachers}
    old_data=json.load(open(BASE_DIR / "data/data.json"))
    data["cal_ids"]=old_data["cal_ids"]
    with open(BASE_DIR / "data/data.json", "w") as f:
        json.dump(data, f)


read_xlsx(BASE_DIR / "data-sources/24-25/Class Schedule of Autumn 2024-25 2nd Sem.xlsx")
