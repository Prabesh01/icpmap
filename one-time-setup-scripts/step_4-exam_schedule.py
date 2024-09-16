from datetime import datetime
import json

import openpyxl

from config import get_basedir

BASE_DIR=get_basedir()
assestment_file = BASE_DIR / "data-sources/23-24/Assessment_Report.xlsx"
exam_data_file = BASE_DIR / "data/exam.json"

exam_data = {}


def parse_assestment_file(source):
    dataframe = openpyxl.load_workbook(source)
    for sheet in dataframe.sheetnames:
        print(sheet)
        exam_data[sheet] = []
        dataframe1=dataframe[sheet]
        module=None
        
        cont=False
        date_indx=0
        mod_idx=0
        code_idx=0
        credit_idx=0
        per_idx=0
        title_idx=0
        
        for row in range(1,dataframe1.max_row):
            col_1=dataframe1[row][1].value
            if not cont and not col_1: continue
            if not cont and 'title' in col_1.lower():
                for i in range(1, dataframe1.max_column):
                    val=str(dataframe1[row][i].value).lower()
                    if 'date' in val: date_indx=i
                    if 'title' in val: mod_idx=i
                    if 'code' in val: code_idx=i
                    if 'credit' in val: credit_idx=i
                    if 'percentage' in val: per_idx=i
                    if 'component' in val: title_idx=i
                if not date_indx or not mod_idx or not title_idx: break
                cont=True
                continue
            if not cont: continue
            if col_1: module = col_1.strip()
            if module and not any(x['module']==module for x in exam_data[sheet]):
                code=dataframe1[row][code_idx].value.strip() if code_idx else ""
                credit=str(dataframe1[row][credit_idx].value).strip() if credit_idx else 0
                exam_data[sheet].append({"module": module, "code": code, "credits": credit, "exams": []})
            exams_index = next((i for i, x in enumerate(exam_data[sheet]) if x['module'] == module), None)
            exam_date_value=dataframe1[row][date_indx].value
            if type(exam_date_value)==str:
                exam_date_value=exam_date_value.strip()
                print(exam_date_value)
                try:
                    exam_dates=[x.split(',')[-1].strip() for x in exam_date_value.split('-')]
                    print(exam_dates)
                    if len(exam_dates[0].split())==4:
                        exam_year=exam_dates[0].split()[-1]
                        zero_split=exam_dates[0].split()
                        exam_dates[0]=zero_split[1]+' '+zero_split[2]
                        if len(exam_dates)==2:
                            if len(exam_dates[1].split())==4:
                                one_split=exam_dates[1].split()
                                exam_dates[1]=one_split[1]+' '+one_split[2]+' '+one_split[3]
                    if len(exam_dates)==1:
                        exam_year=exam_dates[0][-4:].strip()
                        exam_dates[0]=exam_dates[0][:-4].strip()
                        exam_dates.append(exam_dates[0])
                    elif len(exam_dates[1].split())==2:
                        exam_year=datetime.now().year
                        if exam_dates[1][-4]=='2' and exam_dates[1][-3]=='0':
                            exam_dates[1]=exam_dates[1][:-4].strip()
                    else:
                        exam_dates[1]=exam_dates[1][:-4].strip()
                        exam_year=exam_date_value[-4:].strip()
                    fmts = ["%d %B %Y", "%d %b %Y"]
                    try:
                        start_date=datetime.strptime(f"{exam_dates[0]} {exam_year}", fmts[0]).strftime("%Y-%m-%d")
                    except:
                        start_date=datetime.strptime(f"{exam_dates[0]} {exam_year}", fmts[1]).strftime("%Y-%m-%d")
                    try:
                        end_date=datetime.strptime(f"{exam_dates[1]} {exam_year}", fmts[0]).strftime("%Y-%m-%d")
                    except:
                        end_date=datetime.strptime(f"{exam_dates[1]} {exam_year}", fmts[1]).strftime("%Y-%m-%d")                        
                except:
                    exam_dates=exam_date_value.split(',')[-1].split('-')
                    print(exam_dates)
                    exam_dates[1]=exam_dates[1][:-4].strip()
                    exam_year=exam_date_value.split()[-1].strip()
                    start_date=datetime.strptime(f"{exam_dates[0]} {exam_year}", "%d %B %Y").strftime("%Y-%m-%d")
                    end_date=datetime.strptime(f"{exam_dates[1]} {exam_year}", "%d %B %Y").strftime("%Y-%m-%d")                    
            else:
                if not exam_date_value: continue
                start_date=end_date=exam_date_value.strftime("%Y-%m-%d")
            if per_idx:
                percentage=dataframe1[row][per_idx].value
                if not percentage: percentage=100
                else:
                    percentage=int(percentage*100)
            else: percentage=100
            exam_data[sheet][exams_index]['exams'].append({
                'title':dataframe1[row][title_idx].value.strip(),
                'percentage': percentage,
                'start_date': start_date,
                'end_date': end_date
                })
    with open(exam_data_file, 'w') as f:
        json.dump(exam_data, f, indent=4)

parse_assestment_file(assestment_file)
